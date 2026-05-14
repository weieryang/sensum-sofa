import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

data = [
    {'序号': 1, '国家': 'Morocco', '行业': 'Furniture Retail/Wholesale', '公司名': 'Mobilia', '官网': 'https://mobilia.ma', '联系人': 'Yasser Ferssiwi', '职位': 'Directeur des Achats & Supply Chain', '邮箱': 'mobilia@mobilia.ma', '电话': '+212 522-260414', '主营': 'Designer furniture, modular sofas', '采购规模': 'High (Active China sourcing)', '备注': '51-200 employees. Major retail network.'},
    {'序号': 2, '国家': 'Morocco', '行业': 'Furniture Retail', '公司名': 'Poltronesofà Maroc', '官网': 'https://www.poltronesofa.com/fr-MA', '联系人': 'Directeur Réseau', '职位': 'Sourcing & Operations Manager', '邮箱': 'maroc@poltronesofa.com', '电话': '+212 522-943333', '主营': 'Sofas, sofa-beds', '采购规模': 'Medium-High', '备注': '50-150 employees. Specialized in sofas/compact beds.'},
    {'序号': 3, '国家': 'Morocco', '行业': 'Furniture Manufacture/Import', '公司名': 'Dolly Maroc', '官网': 'http://www.dollymaroc.ma', '联系人': 'Responsable Approvisionnement', '职位': 'Purchasing Manager', '邮箱': 'dolly@dollymaroc.ma', '电话': '+212 522-333333', '主营': 'Furniture sets, foam, mattresses', '采购规模': 'Medium-High', '备注': '100-200 employees. Sources chemicals/fabrics from China.'},
    {'序号': 4, '国家': 'Morocco', '行业': 'Furniture Wholesale', '公司名': 'Kaoba', '官网': 'https://www.kaoba.ma', '联系人': 'Responsable Achats', '职位': 'Purchasing Manager', '邮箱': 'contact@kaoba.ma', '电话': '+212 522-231946', '主营': 'Modern decor, living rooms', '采购规模': 'Medium', '备注': '60 employees. Focus on space-saving designs.'},
    {'序号': 5, '国家': 'Morocco', '行业': 'Furniture Import', '公司名': "Interior's Maroc", '官网': 'http://www.interiors.ma', '联系人': 'Directeur Achats', '职位': 'Purchasing Director', '邮箱': 'infos@interiors.ma', '电话': '+212 522-949494', '主营': 'European style, solid wood, compact sets', '采购规模': 'Small-Medium', '备注': '80 employees. Premium importer.'},
    {'序号': 6, '国家': 'Morocco', '行业': 'Furniture Design/Sourcing', '公司名': 'Home Design', '官网': 'http://www.homedesign.ma', '联系人': 'Responsable Sourcing', '职位': 'Sourcing Manager', '邮箱': 'contact@homedesign.ma', '电话': '+212 522-363636', '主营': 'Modular furniture, space-saving living', '采购规模': 'Small-Medium', '备注': '75 employees. Focus on space-optimized furniture.'},
    {'序号': 7, '国家': 'Morocco', '行业': 'Furniture Manufacture/Import', '公司名': 'Socheal (Rêve Bleu)', '官网': 'http://www.revebleu.ma', '联系人': 'Responsable Logistique & Achats', '职位': 'Sourcing Manager', '邮箱': 'contact@socheal.ma', '电话': '+212 522-621960', '主营': 'Bedding, living room furniture', '采购规模': 'High', '备注': 'Major manufacturer (>200 but key importer).'},
    {'序号': 8, '国家': 'Morocco', '行业': 'Furniture Design/Production', '公司名': 'Mazarin', '官网': 'https://mazarin.ma', '联系人': 'Directeur de Production', '职位': 'Production & Sourcing Manager', '邮箱': 'contact@mazarin.ma', '电话': '+212 522-666611', '主营': 'Bespoke furniture, high-end decor', '采购规模': 'Small-Medium', '备注': '50-100 employees. Focus on custom designs.'}
]

file_name = 'Morocco_Compact_Sofa_Leads.xlsx'
df = pd.DataFrame(data)

with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Leads')
    workbook = writer.book
    worksheet = writer.sheets['Leads']

    # Header style
    header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    header_font = Font(bold=True)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

print(f"File {file_name} created successfully.")

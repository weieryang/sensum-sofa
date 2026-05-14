import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

data = [
    {
        '序号': 1, '国家': 'Australia', '行业': 'Furniture Wholesale', '公司名': 'Lounge Lovers', '官网': 'https://www.loungelovers.com.au',
        '联系人': 'James Croft', '职位': 'General Manager', '邮箱': 'info@loungelovers.com.au',
        '电话': '+61 1300 738 088', '主营': 'Sofas, sofa beds, living room furniture',
        '采购规模': 'Large (~2,305 shipments)', '备注': '150-200 employees. Focus on compact designs.'
    },
    {
        '序号': 2, '国家': 'Australia', '行业': 'Furniture Wholesale', '公司名': 'GlobeWest', '官网': 'https://www.globewest.com.au',
        '联系人': 'Gordana Zabic', '职位': 'Purchasing Manager', '邮箱': 'enquiries@globewest.com.au',
        '电话': '+61 3 9518 1660', '主营': 'Designer furniture and homewares',
        '采购规模': 'Large (Vietnam/China/India)', '备注': '94-110 employees. High-end wholesaler.'
    },
    {
        '序号': 3, '国家': 'Australia', '行业': 'Furniture Retail', '公司名': 'Eureka Furniture', '官网': 'https://www.eurekafurniture.com.au',
        '联系人': 'Melissa Hollis', '职位': 'Merchandise Planning Manager', '邮箱': 'enquiries@eurekafurniture.com.au',
        '电话': '+61 7 3717 7600', '主营': 'Solid timber and living room furniture',
        '采购规模': 'Medium-Large (~557 shipments)', '备注': '100-200 employees. Major retail importer.'
    },
    {
        '序号': 4, '国家': 'Australia', '行业': 'Furniture Retail', '公司名': 'Life Interiors', '官网': 'https://lifeinteriors.com.au',
        '联系人': 'Geoff Shore', '职位': 'Management/Operations', '邮箱': 'support@lifeinteriors.com.au',
        '电话': '+61 1300 132 514', '主营': 'Contemporary living furniture',
        '采购规模': 'Large (~1,432 shipments)', '备注': '50-80 employees. Modern style focus.'
    },
    {
        '序号': 5, '国家': 'Australia', '行业': 'Furniture Wholesale', '公司名': 'Trit House', '官网': 'https://www.trithouse.com.au',
        '联系人': 'Brent Sanders', '职位': 'Trade Sales Manager', '邮箱': 'info@trithouse.com.au',
        '电话': '+61 3 9417 1183', '主营': 'Contemporary designer furniture',
        '采购规模': 'Medium (China/Europe/SE Asia)', '备注': '50-80 employees. High design focus.'
    },
    {
        '序号': 6, '国家': 'Australia', '行业': 'Furniture Wholesale', '公司名': 'Cafe Lighting & Living', '官网': 'https://cafelighting.com.au',
        '联系人': 'Sourcing Manager', '职位': 'Purchasing Team', '邮箱': 'sales@cafelighting.com.au',
        '电话': '+61 2 9545 4666', '主营': 'Designer lighting and furniture',
        '采购规模': 'Medium-Large (China focus)', '备注': '50-70 employees. Heavy sourcing from Guangdong.'
    },
    {
        '序号': 7, '国家': 'Australia', '行业': 'Furniture Design/Wholesale', '公司名': 'MCM House', '官网': 'https://www.mcmhouse.com',
        '联系人': 'Drew Rowling', '职位': 'GM Merchandise & Supply Chain', '邮箱': 'info@mcmhouse.com',
        '电话': '+61 2 9662 2580', '主营': 'Relaxed designer living furniture',
        '采购规模': 'Medium-Large (~59 shipments)', '备注': '50-70 employees. Known for linen sofas.'
    },
    {
        '序号': 8, '国家': 'Australia', '行业': 'Furniture Manufacture/Import', '公司名': 'Mulberry Co', '官网': 'https://mulberryco.com.au',
        '联系人': 'Management Team', '职位': 'Purchasing Manager', '邮箱': 'sofa@mulberrylounges.com.au',
        '电话': '+61 2 9773 8820', '主营': 'Premium sofas and lounges',
        '采购规模': 'Small-Medium (China/SE Asia)', '备注': '50-100 employees. Flexible sizing expert.'
    },
    {
        '序号': 9, '国家': 'Australia', '行业': 'Furniture Wholesale', '公司名': 'Horgans', '官网': 'https://www.horgans.com.au',
        '联系人': 'Sophie Newton', '职位': 'Senior Buyer / Purchasing Manager', '邮箱': 'sales@horgans.com.au',
        '电话': '+61 2 9557 7800', '主营': 'Furniture and homewares',
        '采购规模': 'Medium (China/SE Asia)', '备注': '40-60 employees. Wholesaler.'
    },
    {
        '序号': 10, '国家': 'Australia', '行业': 'Furniture Wholesale', '公司名': 'Vivin Imports', '官网': 'https://www.vivinimports.com.au',
        '联系人': 'Robin Singh', '职位': 'Senior Manager, Supply Chain', '邮箱': 'sales@vivinimports.com.au',
        '电话': '+61 2 9011 8080', '主营': 'Wholesale sofas and chairs',
        '采购规模': 'Very Large (~32,225 shipments)', '备注': '39-50 employees. Massive scale importer.'
    },
    {
        '序号': 11, '国家': 'Australia', '行业': 'Furniture Wholesale', '公司名': 'Alliance Furniture', '官网': 'https://alliancefurniture.com.au',
        '联系人': 'David Trope', '职位': 'Director', '邮箱': 'sales@alliancefurniture.com.au',
        '电话': '+61 2 9554 7144', '主营': 'Designer lounges and outdoor',
        '采购规模': 'Large (China/Vietnam)', '备注': '40 employees. Major B2B wholesaler.'
    },
    {
        '序号': 12, '国家': 'Australia', '行业': 'Furniture Wholesale', '公司名': 'Berton Furniture', '官网': 'https://www.bertonfurniture.com.au',
        '联系人': 'Phillip Clayton', '职位': 'Director / Purchasing', '邮箱': 'sales@bertonfurniture.com.au',
        '电话': '+61 2 4732 2155', '主营': 'Lounge, dining, and bedroom',
        '采购规模': 'Medium-Large (China/Malaysia)', '备注': '20-50 employees. Established wholesaler.'
    }
]

df = pd.DataFrame(data)
file_path = "AU_Compact_Sofa_Leads.xlsx"
df.to_excel(file_path, index=False)

from openpyxl import load_workbook
wb = load_workbook(file_path)
ws = wb.active

header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
header_font = Font(bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
    ws.column_dimensions[column].width = max_length + 4

wb.save(file_path)
print(f"Excel file created at: {file_path}")
